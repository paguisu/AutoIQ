from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(r"D:\AutoIQ")
SOURCE_XLSX = ROOT / "Vera" / "Parámetros Comerciales" / "Parámetros-multi.xlsx"
OUT_JSON = ROOT / "tmp_commercial_params_extract" / "commercial_matrix_data.json"


COMPANY_ALIASES = {
    "ALLIANZ": "Allianz",
    "ATM": "ATM",
    "EXPERTA": "Experta",
    "GALENO": "Galeno",
    "INTEGRITY": "Integrity",
    "MAPFRE": "Mapfre",
    "PROVINCIA": "Provincia",
    "RIVADAVIA": "Rivadavia",
    "SANCOR": "Sancor",
    "SAN CRISTOBAL": "San Cristobal",
    "SMG": "SMG",
    "GALICIA SEGUROS": "Galicia",
    "GALICIA": "Galicia",
    "VICTORIA SEGUROS": "Victoria",
    "VICTORIA": "Victoria",
    "ZURICH": "Zurich",
    "MERCANTIL ANDINA": "Mercantil Andina",
    "PROVIDENCIA": "Providencia",
}

ACTIVE_COMPANIES = {
    "Allianz",
    "ATM",
    "Experta",
    "Mapfre",
    "Provincia",
    "Rivadavia",
    "Sancor",
    "SMG",
    "Victoria",
}

NON_MODULE_FIELDS = {
    "ACTIVA AUTO",
    "ACTIVA MOTO",
    "EMISION ACTIVA",
    "EMISIÓN ACTIVA",
}

TECHNICAL_FIELDS = {
    "COD DE PAS",
    "CODIGO PARA LA COTIZACION",
    "CÓDIGO PARA LA COTIZACIÓN",
    "CODIGO DE COTIZACION DE PRODUCTOR",
    "CÓDIGO DE COTIZACIÓN DE PRODUCTOR",
    "CODIGO DE PRODUCTO",
    "CÓDIGO DE PRODUCTO",
    "CODIGO DE TICKET",
    "CÓDIGO DE TICKET",
    "USUARIO WS",
    "PASSWORD WS",
    "PASSWORD",
    "CONTRASEÑA",
    "USUARIO",
    "NUMERO DE ORG",
    "NÚMERO DE ORG",
    "NUMERO DE ORGANIZADOR",
    "NÚMERO DE ORGANIZADOR",
    "CODIGO PRODUCTO",
    "CÓDIGO PRODUCTO",
}

PLAN_FIELDS = {
    "PLAN",
    "PLANES",
}

FIELD_MAP = {
    "MEDIO DE PAGO": ("Medio de pago", "Medio de pago"),
    "MODO DE PAGO": ("Medio de pago", "Medio de pago"),
    "FORMA DE PAGO": ("Medio de pago", "Medio de pago"),
    "ORDIGEN DEL PAGO": ("Medio de pago", "Origen de pago"),
    "ORIGEN DEL PAGO": ("Medio de pago", "Origen de pago"),
    "INSTITUCION FINANCIERA": ("Medio de pago", "Institucion financiera"),
    "TIPO DE FACTURACION": ("Facturacion/Vigencia", "Tipo de facturacion"),
    "PERIODO DE FACTURACION": ("Facturacion/Vigencia", "Periodo de facturacion"),
    "PERIODO DE ACTUALIZACION": ("Facturacion/Vigencia", "Periodo de actualizacion"),
    "PERIODO DE VIGENCIA": ("Facturacion/Vigencia", "Vigencia"),
    "VIGENCIA": ("Facturacion/Vigencia", "Vigencia"),
    "CUOTAS": ("Facturacion/Vigencia", "Cuotas"),
    "CLAUS AJUSTE": ("Ajustes", "Clausula de ajuste"),
    "CLAUSULAS DE AJUSTE": ("Ajustes", "Clausula de ajuste"),
    "AJUSTE CLAUSULA": ("Ajustes", "Clausula de ajuste"),
    "AJUSTE DE CLAUSULA": ("Ajustes", "Clausula de ajuste"),
    "AJUSTE AUTOM": ("Ajustes", "Ajuste automatico"),
    "AJUSTE AUTOMATICO": ("Ajustes", "Ajuste automatico"),
    "AJUSTE COMERCIAL": ("Ajustes", "Ajuste comercial"),
    "AJUSTE": ("Ajustes", "Ajuste"),
    "AJUSTE DE PRIMA": ("Descuentos/Bonificaciones", "Ajuste de prima"),
    "RECARGO DESCUENTO": ("Descuentos/Bonificaciones", "Recargo/descuento"),
    "DESCUENTO": ("Descuentos/Bonificaciones", "Descuento"),
    "DESCUENTO COMERCIAL": ("Descuentos/Bonificaciones", "Descuento comercial"),
    "DESCUENTO COMERCIAL 2": ("Descuentos/Bonificaciones", "Descuento comercial 2"),
    "% DE BONIFICACION": ("Descuentos/Bonificaciones", "Bonificacion"),
    "BONIFICACI%": ("Descuentos/Bonificaciones", "Bonificacion"),
    "BONIFICACION": ("Descuentos/Bonificaciones", "Bonificacion"),
    "BONIFICACION ADICIONAL POR PROMOCION": ("Descuentos/Bonificaciones", "Bonificacion adicional por promocion"),
    "DESCUENTO POR SEG NUEVO": ("Descuentos/Bonificaciones", "Descuento por seguro nuevo"),
    "DTO POR NO SINIESTRO": ("Descuentos/Bonificaciones", "Descuento por no siniestro"),
    "% DE DESC ESPECIAL": ("Descuentos/Bonificaciones", "Porcentaje de descuento especial"),
    "ALTERNATIVA COMERCIAL": ("Descuentos/Bonificaciones", "Alternativa comercial"),
    "BENEFICIOS": ("Descuentos/Bonificaciones", "Beneficios"),
    "CAMPAÑAS": ("Descuentos/Bonificaciones", "Campañas"),
    "RECARGO ADMINISTRATIVO": ("Descuentos/Bonificaciones", "Recargo administrativo"),
    "VARIACION 32080": ("Ajustes", "Variacion 32080"),
    "COEFICIENTE DE RC": ("Ajustes", "Coeficiente RC"),
    "COEFICIENTE RESPONSABILIDAD CIVIL": ("Ajustes", "Coeficiente RC"),
    "COEFICIENTE CASCO": ("Ajustes", "Coeficiente Casco"),
    "AJUSTE RC": ("Ajustes", "Ajuste RC"),
    "AJUSTE CASCO": ("Ajustes", "Ajuste Casco"),
    "RASTREO": ("Rastreo", "Rastreo"),
    "TIENE LOCALIZADOR": ("Rastreo", "Localizador"),
    "LOCALIZADORES": ("Rastreo", "Localizador"),
    "ALARMA": ("Rastreo", "Alarma/rastreo"),
    "COMISION": ("Comisiones", "Comision"),
    "COMISION VARIABLE": ("Comisiones", "Comision variable"),
    "USO": ("Otros comerciales", "Uso"),
    "USOS": ("Otros comerciales", "Uso"),
    "USO SOCIAL": ("Otros comerciales", "Uso social"),
    "GARAGE": ("Otros comerciales", "Garage"),
    "GUARDADO EN GARAGE": ("Otros comerciales", "Garage"),
    "ASISTENCIA MECANICA": ("Otros comerciales", "Asistencia mecanica"),
    "SUMA GRANIZO": ("Otros comerciales", "Suma granizo"),
    "EQUPO DE GNC": ("Otros comerciales", "Equipo GNC"),
    "EQUIPO DE GNC": ("Otros comerciales", "Equipo GNC"),
    "RUEDAS AUXILIARES": ("Otros comerciales", "Ruedas auxiliares"),
    "TIPOS": ("Otros comerciales", "Tipo vehiculo"),
    "CLASES": ("Otros comerciales", "Clase"),
    "CATEGORIA": ("Otros comerciales", "Categoria"),
    "RC": ("Otros comerciales", "RC"),
    "KMS ANUALES": ("Otros comerciales", "Kms anuales"),
    "PORCENTAJE VALOR AUTO": ("Otros comerciales", "Porcentaje valor auto"),
    "LOCALIDAD": ("Otros comerciales", "Localidad"),
}


def norm(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip()


def norm_key(value) -> str:
    value = norm(value).upper()
    value = value.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def unique_headers(headers: list[str]) -> list[str]:
    counts = {}
    result = []
    for header in headers:
        key = norm_key(header)
        counts[key] = counts.get(key, 0) + 1
        if counts[key] == 1:
            result.append(header)
            continue
        if key == "USUARIO":
            result.append("USUARIO WS")
        elif key == "PASSWORD":
            result.append("PASSWORD WS")
        else:
            result.append(f"{header} {counts[key]}")
    return result


def canonical_company(value) -> str:
    key = norm_key(value)
    return COMPANY_ALIASES.get(key, norm(value).title())


def classify(field: str):
    key = norm_key(field)
    if not key:
        return None
    if key in NON_MODULE_FIELDS:
        return ("Fuera del modulo", field)
    if key in TECHNICAL_FIELDS:
        return ("Tecnicos excluidos", field)
    if key in PLAN_FIELDS:
        return ("Tecnicos excluidos", field)
    return FIELD_MAP.get(key, ("Especificos por compania", field))


def parse_blocks(ws):
    blocks = []
    rows = list(ws.iter_rows(values_only=True))
    for row_idx, row in enumerate(rows, start=1):
        cells = [norm(v) for v in row]
        for col_idx, cell in enumerate(cells, start=1):
            if norm_key(cell) != "USUARIO":
                continue
            headers = []
            j = col_idx - 1
            while j < len(cells) and norm(cells[j]):
                headers.append(norm(cells[j]))
                j += 1
            headers = unique_headers(headers)
            if "ASEGURADORA" not in [norm_key(h) for h in headers]:
                continue
            data_rows = []
            r = row_idx + 1
            while r <= len(rows):
                vals = [norm(v) for v in rows[r - 1][col_idx - 1 : col_idx - 1 + len(headers)]]
                if not any(vals):
                    break
                if norm_key(vals[0]) == "USUARIO":
                    break
                if vals[0] and vals[1] and norm_key(vals[1]) != "ASEGURADORA":
                    data_rows.append(dict(zip(headers, vals)))
                r += 1
            if data_rows:
                blocks.append({"sheet": ws.title, "row": row_idx, "col": col_idx, "headers": headers, "rows": data_rows})
    return blocks


def build():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    blocks = parse_blocks(wb["feb 26"])

    companies = []
    company_seen = set()
    default_by_company = {}
    user_rows = []
    excluded_rows = []

    for block in blocks:
        headers = block["headers"]
        for row in block["rows"]:
            company = canonical_company(row.get("ASEGURADORA"))
            if not company:
                continue
            if company not in company_seen:
                company_seen.add(company)
                companies.append(
                    {
                        "company": company,
                        "active": company in ACTIVE_COMPANIES,
                    }
                )

            user = row.get("USUARIO", "")
            record = {
                "usuario": "Default Seguros911" if norm_key(user) == "DEFECTO" else user,
                "company": company,
                "source_sheet": block["sheet"],
                "source_row": block["row"],
                "values": {},
                "excluded": {},
            }
            for header in headers:
                hkey = norm_key(header)
                if hkey in {"USUARIO", "ASEGURADORA"}:
                    continue
                value = row.get(header, "")
                if value == "":
                    continue
                section, concept = classify(header)
                if section in {"Tecnicos excluidos", "Fuera del modulo"}:
                    excluded_rows.append(
                        {
                            "usuario": record["usuario"],
                            "company": company,
                            "section": section,
                            "field": header,
                            "value": value,
                            "source": f"{block['sheet']}!R{block['row']}",
                        }
                    )
                    record["excluded"][header] = value
                    continue
                key = f"{section}||{concept}"
                record["values"][key] = value

            user_rows.append(record)
            if record["usuario"] == "Default Seguros911":
                default_by_company[company] = record["values"]

    concept_order = []
    seen_concepts = set()
    section_rank = {
        "Facturacion/Vigencia": 1,
        "Medio de pago": 2,
        "Ajustes": 3,
        "Descuentos/Bonificaciones": 4,
        "Rastreo": 5,
        "Comisiones": 6,
        "Otros comerciales": 7,
        "Especificos por compania": 8,
    }
    for record in user_rows:
        for key in record["values"]:
            if key not in seen_concepts:
                seen_concepts.add(key)
                section, concept = key.split("||", 1)
                concept_order.append({"section": section, "concept": concept, "key": key})
    concept_order.sort(key=lambda item: (section_rank.get(item["section"], 99), item["concept"]))

    active_company_names = [c["company"] for c in companies if c["active"]]
    inactive_company_names = [c["company"] for c in companies if not c["active"]]
    companies_sorted = [{"company": c, "active": True} for c in sorted(active_company_names)] + [
        {"company": c, "active": False} for c in sorted(inactive_company_names)
    ]

    result = {
        "source": str(SOURCE_XLSX),
        "companies": companies_sorted,
        "concepts": concept_order,
        "defaults": default_by_company,
        "users": user_rows,
        "excluded": excluded_rows,
    }
    OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUT_JSON.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(OUT_JSON)
    print(f"companies={len(companies_sorted)} concepts={len(concept_order)} users={len(user_rows)} excluded={len(excluded_rows)}")


if __name__ == "__main__":
    build()
